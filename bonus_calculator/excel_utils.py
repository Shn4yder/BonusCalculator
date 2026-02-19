import openpyxl
import re
from .utils import normalize_resource_name, to_number

def load_bonuses_from_excel(xlsx_path: str) -> tuple[float | None, float | None]:
    if openpyxl is None:
        raise RuntimeError("Для чтения Excel требуется пакет openpyxl. Установите: pip install openpyxl")
    
    target_col_label = normalize_resource_name("общий итог")
    row_staff_label = normalize_resource_name("премия исполнителей")
    row_manager_label = normalize_resource_name("премия руководителя")
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    staff_bonus = None
    manager_bonus = None
    
    for ws in wb.worksheets:
        positions = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.strip():
                    key = normalize_resource_name(val)
                    if key and key not in positions:
                        positions[key] = cell.coordinate
                        
        if target_col_label in positions:
            col = openpyxl.utils.cell.column_index_from_string(
                re.sub(r"\d+", "", positions[target_col_label])
            )
            
            if row_staff_label in positions and staff_bonus is None:
                row_idx = int(re.sub(r"\D+", "", positions[row_staff_label]))
                v = ws.cell(row=row_idx, column=col).value
                staff_bonus = to_number(v)
                
            if row_manager_label in positions and manager_bonus is None:
                row_idx = int(re.sub(r"\D+", "", positions[row_manager_label]))
                v = ws.cell(row=row_idx, column=col).value
                manager_bonus = to_number(v)
                
        if staff_bonus is not None and manager_bonus is not None:
            break
            
    return staff_bonus, manager_bonus
