import os
import shutil
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from .utils import RUSSIAN_MONTHS

def generate_report(
    report_path: str,
    project_name: str,
    sorted_months: list[tuple[int, int]],
    all_resources: list[tuple[str, object]],
    res_data: dict[int, dict[tuple[int, int], float]],
    staff_bonus: float | None,
    manager_bonus: float | None,
    visible_indices: list[int] | None = None
):
    # Ensure directory exists
    parent_dir = os.path.dirname(report_path)
    if parent_dir and not os.path.exists(parent_dir):
        try:
            os.makedirs(parent_dir, exist_ok=True)
        except Exception:
            pass

    # Load template or create new
    template_path = os.path.join(os.path.dirname(__file__), "..", "data", "templates", "output_data_temp.xlsx")
    template_path = os.path.abspath(template_path)
    
    wb = None
    if os.path.exists(report_path):
        try:
            wb = openpyxl.load_workbook(report_path)
        except Exception:
            pass
            
    if wb is None:
        if os.path.isfile(template_path):
            try:
                shutil.copyfile(template_path, report_path)
                wb = openpyxl.load_workbook(report_path)
            except Exception:
                wb = openpyxl.Workbook()
        else:
            wb = openpyxl.Workbook()
        
    ws = wb.active

    # Находим строку заголовков (обычно строка 2 или 3, будем искать "ФИО" или "Ресурс")
    header_row = 2
    fio_col = 2  # B
    
    # Поиск заголовков в шаблоне
    found_header = False
    for r_idx in range(1, 10):
        for c_idx in range(1, 10):
            val = ws.cell(row=r_idx, column=c_idx).value
            if val and isinstance(val, str) and ("ФИО" in val or "Ресурс" in val):
                header_row = r_idx
                fio_col = c_idx
                found_header = True
                break
        if found_header:
            break
            
    # Если не нашли, предполагаем стандартную структуру
    if not found_header:
        header_row = 3
        fio_col = 2
    
    # Заполняем месяцы в заголовке
    month_start_col = fio_col + 1
    
    # Очистка области заголовков от объединений и старых данных
    total_cols_needed = len(sorted_months) + 3
    # Use max_column to ensure we clear any old headers that might be far to the right
    max_col_needed = max(month_start_col + total_cols_needed, ws.max_column + 1)
    
    ranges_to_unmerge = []
    for rng in ws.merged_cells.ranges:
        if rng.max_col >= month_start_col:
            ranges_to_unmerge.append(rng)
            
    for rng in list(ranges_to_unmerge):
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass

    # Очищаем область заголовков (включая верхние строки, где могут быть объединенные КТУ/Премия)
    # Начинаем с 1-й строки, так как КТУ/Премия пишутся в row=1
    start_header_clear = 1
    for r in range(start_header_clear, header_row + 1):
        for c in range(month_start_col, max_col_needed + 5):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).border = Border() # Also clear borders
    
    # Заполняем месяцы и годы
    current_year_start_col = None
    current_year = None
    
    for i, (y, m) in enumerate(sorted_months):
        col_idx = month_start_col + i
        
        # Год (в строке над месяцем)
        if header_row > 1:
            y_cell = ws.cell(row=header_row-1, column=col_idx)
            y_cell.value = y
            y_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Логика объединения годов
            if current_year is None:
                current_year = y
                current_year_start_col = col_idx
            elif y != current_year:
                # Объединяем предыдущий год
                if current_year_start_col < col_idx - 1:
                    ws.merge_cells(start_row=header_row-1, start_column=current_year_start_col, 
                                   end_row=header_row-1, end_column=col_idx-1)
                current_year = y
                current_year_start_col = col_idx
        
        # Месяц
        month_name = RUSSIAN_MONTHS.get(m, "")
        c = ws.cell(row=header_row, column=col_idx)
        c.value = month_name
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Объединяем последний год
    if header_row > 1 and current_year is not None:
        last_col = month_start_col + len(sorted_months) - 1
        if current_year_start_col < last_col:
             ws.merge_cells(start_row=header_row-1, start_column=current_year_start_col, 
                            end_row=header_row-1, end_column=last_col)

    # Столбец "Итого"
    total_work_col = month_start_col + len(sorted_months)
    
    # Столбец "КТУ"
    ktu_col = total_work_col + 1
    
    # Столбец "Премия"
    bonus_col = ktu_col + 1
    
    # Заголовки Итого/КТУ/Премия
    if header_row > 1:
        # Общий итог (остается под заголовком проекта, занимает 2 строки: Год и Месяц)
        ws.cell(row=header_row-1, column=total_work_col, value="Общий итог")
        ws.merge_cells(start_row=header_row-1, start_column=total_work_col, end_row=header_row, end_column=total_work_col)
        c_tot = ws.cell(row=header_row-1, column=total_work_col)
        c_tot.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_tot.font = Font(bold=False, italic=False)
        
        # КТУ (объединяется с 1-й строки до строки заголовков)
        ws.cell(row=1, column=ktu_col, value="КТУ")
        ws.merge_cells(start_row=1, start_column=ktu_col, end_row=header_row, end_column=ktu_col)
        c_ktu_h = ws.cell(row=1, column=ktu_col)
        c_ktu_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_ktu_h.font = Font(bold=False, italic=False)
        
        # Премия (объединяется с 1-й строки до строки заголовков)
        ws.cell(row=1, column=bonus_col, value="Премия")
        ws.merge_cells(start_row=1, start_column=bonus_col, end_row=header_row, end_column=bonus_col)
        c_bon_h = ws.cell(row=1, column=bonus_col)
        c_bon_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_bon_h.font = Font(bold=True)
    else:
        ws.cell(row=header_row, column=total_work_col, value="Общий итог")
        ws.cell(row=header_row, column=ktu_col, value="КТУ")
        ws.cell(row=header_row, column=bonus_col, value="Премия")

    # -- MERGING A1 --
    ws["A1"] = f'Расчет премии исполнителей по проекту\n"{project_name}"'
         
    try:
        # Объединяем A1 до столбца ПЕРЕД КТУ (то есть до total_work_col включительно)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_work_col)
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws["A1"].font = Font(bold=True)
    except Exception:
        pass

    # ====================================================================================
    # CALCULATION PHASE (Calculate KTU and Bonus ONLY for VISIBLE resources)
    # ====================================================================================
    
    if visible_indices is None:
        visible_indices = list(range(len(all_resources)))
    
    total_hours_map = {} # idx -> hours
    grand_total_hours = 0.0
    
    # Calculate total hours only for visible resources
    for idx in visible_indices:
        hours = sum(res_data.get(idx, {}).values())
        total_hours_map[idx] = hours
        grand_total_hours += hours
        
    # Identify max work resource among visible ones for plug logic
    max_work_idx = -1
    max_val = -1.0
    
    if grand_total_hours > 0 and len(visible_indices) > 0:
        # Initialize with first visible
        first_idx = visible_indices[0]
        max_work_idx = first_idx
        max_val = total_hours_map.get(first_idx, 0.0)
        
        for idx in visible_indices:
            val = total_hours_map.get(idx, 0.0)
            if val > max_val:
                max_val = val
                max_work_idx = idx
                
    ktu_map = {}
    bonus_map = {}
    
    current_ktu_sum = 0.0
    current_bonus_sum = 0.0
    
    # Calculate initial values for visible resources (skipping max work resource for now)
    for idx in visible_indices:
        if idx == max_work_idx:
            continue
            
        h = total_hours_map.get(idx, 0.0)
        if grand_total_hours > 0:
            ktu = round(h * 100 / grand_total_hours, 2)
        else:
            ktu = 0.0
        ktu_map[idx] = ktu
        current_ktu_sum += ktu
        
        if staff_bonus is not None:
             bon = round(ktu / 100 * staff_bonus, 2)
             bonus_map[idx] = bon
             current_bonus_sum += bon
        else:
             bonus_map[idx] = 0.0
             
    # Apply plug logic to max work resource
    if max_work_idx != -1:
        ktu_map[max_work_idx] = round(100.0 - current_ktu_sum, 2)
        if staff_bonus is not None:
            bonus_map[max_work_idx] = round(staff_bonus - current_bonus_sum, 2)
        else:
            bonus_map[max_work_idx] = 0.0
    
    # ====================================================================================
    # PREPARE DATA WRITING
    # ====================================================================================

    start_data_row = header_row + 1
    
    # 1. CLEANUP OLD DATA
    # Strategy: Find the start of the Footer (Signatures).
    # Everything between start_data_row and Footer is considered "Old Data" and must be deleted.
    # This ensures we remove not just the previous run's data, but also any "stuck" blocks 
    # like the one the user highlighted (rows 10-14 in their screenshot).
    
    signature_keywords = [
        "руководитель проекта", 
        "куратор", 
        "заказчик", 
        "инвестор",
        "согласовано",
        "утверждаю",
        "главный конструктор"
    ]
    
    footer_start_row = None
    
    # Search for signature in Column A and FIO Column (B)
    # Scan a reasonable range (e.g. 1000 rows)
    for r in range(start_data_row, start_data_row + 1000):
        # Check FIO Col
        val_fio = ws.cell(row=r, column=fio_col).value
        str_fio = str(val_fio).strip().lower() if val_fio else ""
        
        # Check Col A (sometimes signatures are in A)
        val_a = ws.cell(row=r, column=1).value
        str_a = str(val_a).strip().lower() if val_a else ""
        
        found_sig = False
        for kw in signature_keywords:
            if kw in str_fio or kw in str_a:
                found_sig = True
                break
        
        if found_sig:
            footer_start_row = r
            break
            
    if footer_start_row:
        # Delete everything from start_data_row up to (but not including) the signature row
        rows_to_delete = footer_start_row - start_data_row
        if rows_to_delete > 0:
            ws.delete_rows(start_data_row, amount=rows_to_delete)
    else:
        # Fallback: If no signatures found, try to find the LAST "Total" or "Manager Bonus"
        # and delete everything up to it?
        # Or just delete to max_row if it's reasonable?
        
        # If we can't find signatures, it's safer to clear everything below header
        # assuming the file only contains this report.
        max_r = ws.max_row
        if max_r >= start_data_row:
            ws.delete_rows(start_data_row, amount=max_r - start_data_row + 1)

    # 2. INSERT NEW ROWS
    # +1 for "Total" row
    rows_needed = len(visible_indices) + 1 
    ws.insert_rows(start_data_row, amount=rows_needed)

    # 3. WRITE DATA ROWS
    current_row_offset = 0
    
    for idx in visible_indices:
        if idx >= len(all_resources):
            continue
            
        name = all_resources[idx][0]
        monthly_data = res_data.get(idx, {})
        
        row = start_data_row + current_row_offset
        current_row_offset += 1
        
        ws.cell(row=row, column=1, value=current_row_offset)
        ws.cell(row=row, column=fio_col, value=name)
        
        # Hours
        for i, (y, m) in enumerate(sorted_months):
            col_idx = month_start_col + i
            hours = monthly_data.get((y, m), 0.0)
            c_h = ws.cell(row=row, column=col_idx, value=hours)
            c_h.number_format = '0.00'
            c_h.alignment = Alignment(horizontal='center', vertical='center')
            c_h.font = Font(bold=False, italic=False)
            
        # Row Total (Formula is safe as it sums visible cells in the row)
        start_col_letter = openpyxl.utils.get_column_letter(month_start_col)
        end_col_letter = openpyxl.utils.get_column_letter(total_work_col - 1)
        c_sum = ws.cell(row=row, column=total_work_col, value=f"=ROUND(SUM({start_col_letter}{row}:{end_col_letter}{row}), 2)")
        c_sum.number_format = '0.00'
        c_sum.alignment = Alignment(horizontal='center', vertical='center')
        c_sum.font = Font(bold=False, italic=False)
        
        # KTU (Calculated Value)
        c_ktu = ws.cell(row=row, column=ktu_col, value=ktu_map.get(idx, 0.0))
        c_ktu.number_format = '0.00'
        c_ktu.alignment = Alignment(horizontal='center', vertical='center')
        c_ktu.font = Font(bold=False, italic=False)
        
        # Bonus (Calculated Value)
        c_bon = ws.cell(row=row, column=bonus_col, value=bonus_map.get(idx, 0.0))
        c_bon.number_format = '0.00'
        c_bon.alignment = Alignment(horizontal='center', vertical='center')
        c_bon.font = Font(bold=True)

    # 4. WRITE TOTAL ROW
    total_row_idx = start_data_row + len(visible_indices)
    
    c_total_label = ws.cell(row=total_row_idx, column=fio_col, value="Общий итог")
    c_total_label.font = Font(bold=False, italic=False)
    c_total_label.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sums (using formulas to sum visible rows)
    for i in range(len(sorted_months)):
        col_idx = month_start_col + i
        col_let = openpyxl.utils.get_column_letter(col_idx)
        c_tot_m = ws.cell(row=total_row_idx, column=col_idx, value=f"=ROUND(SUM({col_let}{start_data_row}:{col_let}{total_row_idx-1}), 2)")
        c_tot_m.number_format = '0.00'
        c_tot_m.alignment = Alignment(horizontal='center', vertical='center')
        c_tot_m.font = Font(bold=False, italic=False)
        
    # Total Work Sum
    total_work_col_letter = openpyxl.utils.get_column_letter(total_work_col)
    c_tot_w = ws.cell(row=total_row_idx, column=total_work_col, value=f"=ROUND(SUM({total_work_col_letter}{start_data_row}:{total_work_col_letter}{total_row_idx-1}), 2)")
    c_tot_w.number_format = '0.00'
    c_tot_w.alignment = Alignment(horizontal='center', vertical='center')
    c_tot_w.font = Font(bold=False, italic=False)
    
    # KTU Sum
    ktu_col_letter = openpyxl.utils.get_column_letter(ktu_col)
    c_ktu_tot = ws.cell(row=total_row_idx, column=ktu_col, value=f"=ROUND(SUM({ktu_col_letter}{start_data_row}:{ktu_col_letter}{total_row_idx-1}), 2)")
    c_ktu_tot.number_format = '0.00'
    c_ktu_tot.alignment = Alignment(horizontal='center', vertical='center')
    c_ktu_tot.font = Font(bold=False, italic=False)
    
    # Bonus Sum
    bonus_col_letter = openpyxl.utils.get_column_letter(bonus_col)
    c_bon_tot = ws.cell(row=total_row_idx, column=bonus_col, value=f"=ROUND(SUM({bonus_col_letter}{start_data_row}:{bonus_col_letter}{total_row_idx-1}), 2)")
    c_bon_tot.number_format = '0.00'
    c_bon_tot.alignment = Alignment(horizontal='center', vertical='center')
    c_bon_tot.font = Font(bold=True)

    # 5. WRITE MANAGER BONUS ROW
    # Add empty row between Total and Manager Bonus if needed?
    # Usually we want:
    # ...
    # Total
    # (empty?)
    # Manager Bonus
    
    # If we want an empty row, we should have added +2 to insert_rows.
    # Let's insert rows for spacing AND for the Manager Bonus itself to avoid overwriting footer.
    # We want:
    # Total
    # (Empty)
    # Manager Bonus
    # (Existing Footer starts here)
    
    ws.insert_rows(total_row_idx + 1, amount=2)
    
    manager_row = total_row_idx + 2
    ws.cell(row=manager_row, column=fio_col, value="Премия РП")
    c_m = ws.cell(row=manager_row, column=bonus_col, value=manager_bonus if manager_bonus is not None else 0)
    c_m.number_format = '0.00'
    c_m.alignment = Alignment(horizontal='center', vertical='center')
    c_m.font = Font(bold=True)

    # -- BORDERS --
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    # Main table
    start_border_row = header_row - 1 if header_row > 1 else header_row
    end_border_row = total_row_idx
    
    # Ensure borders for merged headers (KTU/Bonus) which start at row 1
    if start_border_row > 1:
        for r in range(1, start_border_row):
            ws.cell(row=r, column=ktu_col).border = thin_border
            ws.cell(row=r, column=bonus_col).border = thin_border

    for r in range(start_border_row, end_border_row + 1):
        for c in range(1, bonus_col + 1):
            ws.cell(row=r, column=c).border = thin_border
            
    # Manager row - WITH BORDERS
    # Assuming we want to border the cells that have content (Label and Value)
    # The label is at fio_col, the value is at bonus_col
    # If you want a continuous row border, we should iterate from fio_col to bonus_col?
    # Or just border the two cells as they are separate?
    # Usually "Manager Bonus" looks like a separate small table or just two cells.
    # Let's border from fio_col to bonus_col to make it look like a row in the table, 
    # even if intermediate cells are empty.
    
    for c in range(1, bonus_col + 1):
        ws.cell(row=manager_row, column=c).border = thin_border

    # -- COLUMN A FORMATTING --
    exclude_keywords = [
        "премия рп", 
        "заказчик", 
        "куратор", 
        "руководитель", 
        "инвестор"
    ]
    
    calibri_font = Font(name='Calibri', size=11, bold=False)
    left_align = Alignment(horizontal='left', vertical='center')
    
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        cell_a = ws.cell(row=r, column=1)
        val_a = str(cell_a.value).strip().lower() if cell_a.value else ""
        
        should_skip = False
        for kw in exclude_keywords:
            if kw in val_a:
                should_skip = True
                break
        
        if not should_skip:
             cell_a.font = calibri_font
             cell_a.alignment = left_align

    # -- COLUMN WIDTHS --
    # Set fixed widths for columns to prevent excessive stretching
    
    if fio_col > 1:
        # If FIO is not in the first column, assume first column is for numbering
        ws.column_dimensions['A'].width = 5
        fio_col_letter = openpyxl.utils.get_column_letter(fio_col)
        ws.column_dimensions[fio_col_letter].width = 40
    else:
        # If FIO is in the first column (A), set it to wide
        ws.column_dimensions['A'].width = 40
    
    # Month Columns: Standard fixed width
    for i in range(len(sorted_months)):
        col_idx = month_start_col + i
        col_let = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_let].width = 12
        
    # Total, KTU, Bonus Columns: Standard fixed width
    total_col_let = openpyxl.utils.get_column_letter(total_work_col)
    ws.column_dimensions[total_col_let].width = 15
    
    ktu_col_let = openpyxl.utils.get_column_letter(ktu_col)
    ws.column_dimensions[ktu_col_let].width = 10
    
    bonus_col_let = openpyxl.utils.get_column_letter(bonus_col)
    ws.column_dimensions[bonus_col_let].width = 15

    # Сохраняем
    try:
        wb.save(report_path)
        print(f"Отчет успешно сформирован: {report_path}")
    except Exception as e:
        print(f"Ошибка при сохранении отчета: {e}")
