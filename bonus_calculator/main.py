import sys
import os
import re
from collections import OrderedDict
try:
    import aspose.tasks as tsk
except Exception as e:
    print("Не удалось импортировать библиотеку aspose-tasks. Установите пакет: pip install aspose-tasks")
    sys.exit(1)
try:
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font
except Exception:
    openpyxl = None


def normalize_resource_name(name: str) -> str:
    if name is None:
        return ""
    s = str(name).replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s.strip())
    return s.casefold()


def get_resource_name(resource) -> str:
    name = None
    if hasattr(resource, "name"):
        try:
            name = resource.name
        except Exception:
            name = None
    if not name:
        try:
            if hasattr(tsk, "Rsc") and hasattr(resource, "get"):
                name = resource.get(tsk.Rsc.NAME)
        except Exception:
            name = None
    return name if isinstance(name, str) else (str(name) if name is not None else "")


def load_unique_resources(mpp_path: str) -> list[str]:
    project = tsk.Project(mpp_path)
    unique = OrderedDict()
    for r in project.resources:
        raw = get_resource_name(r) or ""
        norm = normalize_resource_name(raw)
        if not norm:
            continue
        if norm not in unique:
            unique[norm] = re.sub(r"\s+", " ", raw.replace("\u00A0", " ").strip())
    return list(unique.values())


def parse_indices(input_str: str, max_index: int) -> list[int]:
    tokens = re.findall(r"\d+", input_str)
    seen = set()
    result = []
    for t in tokens:
        try:
            n = int(t)
            if 1 <= n <= max_index and n not in seen:
                seen.add(n)
                result.append(n)
        except Exception:
            continue
    return result


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    s = s.replace("\u00A0", " ")
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    m = re.search(r"-?\d+(\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def load_bonuses_from_excel(xlsx_path: str) -> tuple[float | None, float | None]:
    if openpyxl is None:
        raise RuntimeError("Для чтения Excel требуется пакет openpyxl. Установите: pip install openpyxl")
    target_col_label = normalize_resource_name("общий итог")
    row_staff_label = normalize_resource_name("премия исполнителей")
    row_manager_label = normalize_resource_name("премия руководителя")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    staff_bonus = None
    manager_bonus = None
    for ws in wb.worksheets:
        positions = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.strip():
                    key = normalize_resource_name(val)
                    if key and key not in positions:
                        positions[key] = cell.coordinate
        if target_col_label in positions:
            col = openpyxl.utils.cell.column_index_from_string(
                re.sub(r"\d+", "", positions[target_col_label])
            )
            if row_staff_label in positions and staff_bonus is None:
                row_idx = int(re.sub(r"\D+", "", positions[row_staff_label]))
                v = ws.cell(row=row_idx, column=col).value
                staff_bonus = _to_number(v)
            if row_manager_label in positions and manager_bonus is None:
                row_idx = int(re.sub(r"\D+", "", positions[row_manager_label]))
                v = ws.cell(row=row_idx, column=col).value
                manager_bonus = _to_number(v)
        if staff_bonus is not None and manager_bonus is not None:
            break
    return staff_bonus, manager_bonus


def main():
    if len(sys.argv) != 4:
        print("Использование: python -m bonus_calculator <путь_к_файлу.mpp> <путь_к_файлу.xlsx> <путь_для_отчета>")
        sys.exit(2)
    mpp_arg = sys.argv[1]
    xlsx_arg = sys.argv[2]
    report_path = sys.argv[3]
    def resolve_file(p: str) -> str:
        p1 = os.path.abspath(p)
        if os.path.isfile(p1):
            return p1
        base = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        p2 = os.path.abspath(os.path.join(base, p))
        if os.path.isfile(p2):
            return p2
        return p1
    mpp_path = resolve_file(mpp_arg)
    xlsx_path = resolve_file(xlsx_arg)
    if not os.path.isfile(mpp_path) or not mpp_path.lower().endswith(".mpp"):
        print("Ошибка: первый аргумент должен быть существующим файлом с расширением .mpp")
        sys.exit(2)
    if not os.path.isfile(xlsx_path) or not xlsx_path.lower().endswith(".xlsx"):
        print("Ошибка: второй аргумент должен быть существующим файлом с расширением .xlsx")
        sys.exit(2)

    # Определяем, является ли report_path путем к файлу или директорией
    final_report_path = report_path
    if report_path.lower().endswith(".xlsx"):
        # Это файл
        parent_dir = os.path.dirname(report_path)
        if parent_dir and not os.path.exists(parent_dir):
            try:
                os.makedirs(parent_dir, exist_ok=True)
            except Exception:
                pass
    else:
        # Это директория
        if not os.path.exists(report_path):
            try:
                os.makedirs(report_path, exist_ok=True)
            except Exception:
                pass
        final_report_path = os.path.join(report_path, "output_data.xlsx")

    try:
        staff_bonus, manager_bonus = load_bonuses_from_excel(xlsx_path)
        staff_bonus, manager_bonus = round(staff_bonus, 2), round(manager_bonus, 2)
        print("Премии из Excel:")
        print(f"Премия исполнителей (общий итог): {staff_bonus}")
        print(f"Премия руководителя (общий итог): {manager_bonus}")
    except Exception as e:
        print(f"Ошибка чтения премий из Excel: {e}")
    try:
        resources = load_unique_resources(mpp_path)
    except Exception as e:
        print(f"Ошибка при чтении .mpp: {e}")
        sys.exit(1)
    if not resources:
        print("Ресурсы не найдены.")
        sys.exit(0)
    for i, r in enumerate(resources, start=1):
        print(f"{i}. {r}")

    # selection = input("Введите номера необходимых ресурсов через запятую: ")
    # indices = parse_indices(selection, len(resources))
    # chosen = [resources[i - 1] for i in indices]
    
    # Теперь используем все ресурсы по умолчанию, как указано в требовании
    chosen = resources
    
    if not chosen:
        print("Не выбрано ни одного ресурса.")
        sys.exit(0)

    # 3. Collect Timephased Data
    import datetime
    from collections import defaultdict

    # Инициализируем проект заново для работы с объектами (ранее мы загружали только имена)
    project = tsk.Project(mpp_path)

    def parse_duration_to_hours(val_str):
        # Простейший парсер для строк вида PT8H0M0S
        if not val_str:
            return 0.0
        try:
            import re
            m = re.search(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(\.\d+)?)S)?', str(val_str))
            if m:
                h = int(m.group(1) or 0)
                mm = int(m.group(2) or 0)
                s = float(m.group(3) or 0)
                return h + mm / 60.0 + s / 3600.0
            return float(val_str)
        except Exception:
            return 0.0

    # Сопоставим выбранные имена с объектами ресурсов
    # Нам нужно найти те же ресурсы, что были выбраны пользователем.
    norm_map = {}
    for r in project.resources:
        raw = get_resource_name(r)
        norm = normalize_resource_name(raw)
        if norm and norm not in norm_map:
            norm_map[norm] = r
            
    selected_resources = []
    for name in chosen:
        norm = normalize_resource_name(name)
        if norm in norm_map:
            selected_resources.append((name, norm_map[norm]))

    res_data = {}
    
    # Определяем диапазон дат проекта для заголовков месяцев
    project_start = project.start_date
    project_finish = project.finish_date
    
    sorted_months = []
    if project_start and project_finish:
        # Приводим к началу месяца для корректного сравнения
        curr = datetime.datetime(project_start.year, project_start.month, 1)
        end_dt = datetime.datetime(project_finish.year, project_finish.month, 1)
        while curr <= end_dt:
            sorted_months.append((curr.year, curr.month))
            if curr.month == 12:
                curr = datetime.datetime(curr.year + 1, 1, 1)
            else:
                curr = datetime.datetime(curr.year, curr.month + 1, 1)
    
    print("Сбор данных о трудозатратах...")
    
    # Map resource UID to our selected resource index
    uid_to_index = {}
    for idx, (name, r) in enumerate(selected_resources):
        try:
            uid = None
            if hasattr(r, "uid"):
                uid = r.uid
            elif hasattr(r, "get"):
                uid = r.get(tsk.Rsc.UID)
            if uid is not None:
                uid_to_index[uid] = idx
        except Exception:
            pass

    # Initialize res_data
    for idx in range(len(selected_resources)):
        res_data[idx] = defaultdict(float)

    # Iterate assignments to aggregate work
    for ra in project.resource_assignments:
        try:
            r = ra.resource
            if r is None:
                continue
            
            r_uid = None
            if hasattr(r, "uid"):
                r_uid = r.uid
            elif hasattr(r, "get"):
                r_uid = r.get(tsk.Rsc.UID)
                
            if r_uid is not None and r_uid in uid_to_index:
                idx = uid_to_index[r_uid]
                
                td_collection = ra.get_timephased_data(project.start_date, project.finish_date, tsk.TimephasedDataType.ASSIGNMENT_WORK)
                
                for td in td_collection:
                    val = td.value
                    hours = parse_duration_to_hours(val)
                    if hours > 0:
                        dt = td.start
                        key = (dt.year, dt.month)
                        res_data[idx][key] += hours
        except Exception as e:
            pass

    # 4. Prepare Excel
    if openpyxl is None:
        print("Ошибка: отсутствует openpyxl.")
        sys.exit(1)
        
    try:
        # Пробуем открыть файл отчета, если он уже есть
        if os.path.exists(final_report_path):
            wb = openpyxl.load_workbook(final_report_path)
        else:
            # Если не найден - пробуем скопировать из шаблона
            template_path = r"C:\Users\a.komarkova\Documents\projects\BonusCalculator\data\templates\output_data.xlsx"
            if os.path.isfile(template_path):
                import shutil
                shutil.copyfile(template_path, final_report_path)
                wb = openpyxl.load_workbook(final_report_path)
            else:
                wb = openpyxl.Workbook()
        ws = wb.active
    except Exception as e:
        print(f"Ошибка при создании/открытии отчета: {e}")
        wb = openpyxl.Workbook()
        ws = wb.active

    # Тема проекта
    project_name = ""
    try:
        # Пробуем получить имя первой задачи (ID 1)
        if project.root_task:
            # Ищем задачу с ID 1 среди дочерних
            task_1 = None
            for child in project.root_task.children:
                if child.id == 1:
                    task_1 = child
                    break
            
            if task_1 and task_1.name:
                project_name = task_1.name
            elif project.root_task.name:
                # Если задачи с ID 1 нет или имя пустое, берем имя корневой задачи
                project_name = project.root_task.name
    except Exception:
        pass
        
    if not project_name:
         # Fallback to Subject/Title
         try:
             if hasattr(project, "subject") and project.subject:
                 project_name = project.subject
             elif hasattr(project, "title") and project.title:
                 project_name = project.title
         except:
             pass
             
    if not project_name:
         project_name = ""

    # Находим строку заголовков (обычно строка 2 или 3, будем искать "ФИО" или "Ресурс")
    header_row = 2
    fio_col = 2  # B
    
    # Поиск заголовков в шаблоне
    found_header = False
    for r_idx in range(1, 10):
        for c_idx in range(1, 10):
            val = ws.cell(row=r_idx, column=c_idx).value
            if val and isinstance(val, str) and ("ФИО" in val or "Ресурс" in val):
                header_row = r_idx
                fio_col = c_idx
                found_header = True
                break
        if found_header:
            break
            
    # Если не нашли, предполагаем стандартную структуру
    if not found_header:
        header_row = 3
        fio_col = 2
    
    # Заполняем месяцы в заголовке
    month_start_col = fio_col + 1
    
    # Очистка области заголовков от объединений и старых данных
    total_cols_needed = len(sorted_months) + 3
    max_col_needed = month_start_col + total_cols_needed
    
    ranges_to_unmerge = []
    for rng in ws.merged_cells.ranges:
        if rng.max_col >= month_start_col:
            ranges_to_unmerge.append(rng)
            
    for rng in list(ranges_to_unmerge):
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass

    # Очищаем две строки заголовков (Год и Месяц)
    start_header_clear = header_row - 1 if header_row > 1 else header_row
    for r in range(start_header_clear, header_row + 1):
        for c in range(month_start_col, max_col_needed + 5):
            ws.cell(row=r, column=c).value = None
    
    # Заполняем месяцы и годы
    current_year_start_col = None
    current_year = None
    
    RUSSIAN_MONTHS = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    
    for i, (y, m) in enumerate(sorted_months):
        col_idx = month_start_col + i
        
        # Год (в строке над месяцем)
        if header_row > 1:
            y_cell = ws.cell(row=header_row-1, column=col_idx)
            y_cell.value = y
            y_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Логика объединения годов
            if current_year is None:
                current_year = y
                current_year_start_col = col_idx
            elif y != current_year:
                # Объединяем предыдущий год
                if current_year_start_col < col_idx - 1:
                    ws.merge_cells(start_row=header_row-1, start_column=current_year_start_col, 
                                   end_row=header_row-1, end_column=col_idx-1)
                current_year = y
                current_year_start_col = col_idx
        
        # Месяц
        month_name = RUSSIAN_MONTHS.get(m, "")
        c = ws.cell(row=header_row, column=col_idx)
        c.value = month_name
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Объединяем последний год
    if header_row > 1 and current_year is not None:
        last_col = month_start_col + len(sorted_months) - 1
        if current_year_start_col < last_col:
             ws.merge_cells(start_row=header_row-1, start_column=current_year_start_col, 
                            end_row=header_row-1, end_column=last_col)

    # Столбец "Итого"
    total_work_col = month_start_col + len(sorted_months)
    
    # Столбец "КТУ"
    ktu_col = total_work_col + 1
    
    # Столбец "Премия"
    bonus_col = ktu_col + 1
    
    # Заголовки Итого/КТУ/Премия
    if header_row > 1:
        # Общий итог (остается под заголовком проекта, занимает 2 строки: Год и Месяц)
        ws.cell(row=header_row-1, column=total_work_col, value="Общий итог")
        ws.merge_cells(start_row=header_row-1, start_column=total_work_col, end_row=header_row, end_column=total_work_col)
        c_tot = ws.cell(row=header_row-1, column=total_work_col)
        c_tot.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_tot.font = Font(bold=False, italic=False)
        
        # КТУ (объединяется с 1-й строки до строки заголовков)
        ws.cell(row=1, column=ktu_col, value="КТУ")
        ws.merge_cells(start_row=1, start_column=ktu_col, end_row=header_row, end_column=ktu_col)
        c_ktu_h = ws.cell(row=1, column=ktu_col)
        c_ktu_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_ktu_h.font = Font(bold=False, italic=False)
        
        # Премия (объединяется с 1-й строки до строки заголовков)
        ws.cell(row=1, column=bonus_col, value="Премия")
        ws.merge_cells(start_row=1, start_column=bonus_col, end_row=header_row, end_column=bonus_col)
        c_bon_h = ws.cell(row=1, column=bonus_col)
        c_bon_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_bon_h.font = Font(bold=True)
    else:
        ws.cell(row=header_row, column=total_work_col, value="Общий итог")
        ws.cell(row=header_row, column=ktu_col, value="КТУ")
        ws.cell(row=header_row, column=bonus_col, value="Премия")

    # -- MERGING A1 --
    cell_a1 = ws["A1"]
    ws["A1"] = f'Расчет премии исполнителей по проекту\n"{project_name}"'
         
    try:
        # Объединяем A1 до столбца ПЕРЕД КТУ (то есть до total_work_col включительно)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_work_col)
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws["A1"].font = Font(bold=True)
    except Exception:
        pass

    # Заполняем данные
    start_data_row = header_row + 1
    
    # Очистка строк данных
    for r in range(start_data_row, start_data_row + 100):
        if ws.cell(row=r, column=fio_col).value is None:
            break
        
    for idx, (name, r_obj) in enumerate(selected_resources):
        row = start_data_row + idx
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=fio_col, value=name)
        
        monthly_data = res_data.get(idx, {})
        
        # Заполняем часы
        for i, (y, m) in enumerate(sorted_months):
            col_idx = month_start_col + i
            hours = monthly_data.get((y, m), 0.0)
            c_h = ws.cell(row=row, column=col_idx, value=hours)
            c_h.number_format = '0.00'
            c_h.alignment = Alignment(horizontal='center', vertical='center')
            c_h.font = Font(bold=False, italic=False)
            
        # Формула Итого
        start_col_letter = openpyxl.utils.get_column_letter(month_start_col)
        end_col_letter = openpyxl.utils.get_column_letter(total_work_col - 1)
        c_sum = ws.cell(row=row, column=total_work_col, value=f"=ROUND(SUM({start_col_letter}{row}:{end_col_letter}{row}), 2)")
        c_sum.number_format = '0.00'
        c_sum.alignment = Alignment(horizontal='center', vertical='center')
        c_sum.font = Font(bold=False, italic=False)

    # Строка ИТОГО
    total_row_idx = start_data_row + len(selected_resources)
    c_total_label = ws.cell(row=total_row_idx, column=fio_col, value="Общий итог")
    c_total_label.font = Font(bold=False, italic=False)
    c_total_label.alignment = Alignment(horizontal='center', vertical='center')
    
    # Суммы по месяцам
    for i in range(len(sorted_months)):
        col_idx = month_start_col + i
        col_let = openpyxl.utils.get_column_letter(col_idx)
        c_tot_m = ws.cell(row=total_row_idx, column=col_idx, value=f"=ROUND(SUM({col_let}{start_data_row}:{col_let}{total_row_idx-1}), 2)")
        c_tot_m.number_format = '0.00'
        c_tot_m.alignment = Alignment(horizontal='center', vertical='center')
        c_tot_m.font = Font(bold=False, italic=False)
        
    # Сумма общих итогов
    total_work_col_letter = openpyxl.utils.get_column_letter(total_work_col)
    c_tot_w = ws.cell(row=total_row_idx, column=total_work_col, value=f"=ROUND(SUM({total_work_col_letter}{start_data_row}:{total_work_col_letter}{total_row_idx-1}), 2)")
    c_tot_w.number_format = '0.00'
    c_tot_w.alignment = Alignment(horizontal='center', vertical='center')
    c_tot_w.font = Font(bold=False, italic=False)
    
    # Сумма КТУ
    ktu_col_letter = openpyxl.utils.get_column_letter(ktu_col)
    c_ktu_tot = ws.cell(row=total_row_idx, column=ktu_col, value=f"=ROUND(SUM({ktu_col_letter}{start_data_row}:{ktu_col_letter}{total_row_idx-1}), 2)")
    c_ktu_tot.number_format = '0.00'
    c_ktu_tot.alignment = Alignment(horizontal='center', vertical='center')
    c_ktu_tot.font = Font(bold=False, italic=False)
    
    # Сумма Премий
    bonus_col_letter = openpyxl.utils.get_column_letter(bonus_col)
    c_bon_tot = ws.cell(row=total_row_idx, column=bonus_col, value=f"=ROUND(SUM({bonus_col_letter}{start_data_row}:{bonus_col_letter}{total_row_idx-1}), 2)")
    c_bon_tot.number_format = '0.00'
    c_bon_tot.alignment = Alignment(horizontal='center', vertical='center')
    c_bon_tot.font = Font(bold=True)

    # КТУ и Премии
    for idx in range(len(selected_resources)):
        row = start_data_row + idx
        
        # КТУ
        c_ktu = ws.cell(row=row, column=ktu_col, value=f"=ROUND({total_work_col_letter}{row}*100/${total_work_col_letter}${total_row_idx}, 2)")
        c_ktu.number_format = '0.00'
        c_ktu.alignment = Alignment(horizontal='center', vertical='center')
        c_ktu.font = Font(bold=False, italic=False)
        
        # Премия
        ktu_col_letter = openpyxl.utils.get_column_letter(ktu_col)
        if staff_bonus is not None:
             c_bon = ws.cell(row=row, column=bonus_col, value=f"=ROUND({ktu_col_letter}{row}/100*{staff_bonus}, 2)")
        else:
             c_bon = ws.cell(row=row, column=bonus_col, value=0)
        c_bon.number_format = '0.00'
        c_bon.alignment = Alignment(horizontal='center', vertical='center')
        c_bon.font = Font(bold=True)

    # Строка "Премия руководителя"
    manager_row = total_row_idx + 2
    ws.cell(row=manager_row, column=fio_col, value="Премия РП")
    c_m = ws.cell(row=manager_row, column=bonus_col, value=manager_bonus if manager_bonus is not None else 0)
    c_m.number_format = '0.00'
    c_m.alignment = Alignment(horizontal='center', vertical='center')
    c_m.font = Font(bold=True)

    # -- BORDERS --
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    # Main table
    start_border_row = header_row - 1 if header_row > 1 else header_row
    end_border_row = total_row_idx
    
    for r in range(start_border_row, end_border_row + 1):
        for c in range(1, bonus_col + 1):
            ws.cell(row=r, column=c).border = thin_border
            
    # Manager row - NO BORDERS
    # Explicitly removing borders for manager row (just in case they were inherited)
    ws.cell(row=manager_row, column=fio_col).border = None
    ws.cell(row=manager_row, column=bonus_col).border = None

    # Сохраняем
    try:
        wb.save(final_report_path)
        print(f"Отчет успешно сформирован: {final_report_path}")
    except Exception as e:
        print(f"Ошибка при сохранении отчета: {e}")


if __name__ == "__main__":
    main()

